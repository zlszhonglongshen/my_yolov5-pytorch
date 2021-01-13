from torch.utils.data import DataLoader
from torch.optim import SGD, Adam, lr_scheduler
import argparse
from dataset import MyDataset_xml, MyDataset_json
import torch
import numpy as np
from tqdm import tqdm
import os,shutil
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from loss_fun import compute_loss, hyp
from main_test import evaluate
from yolov5s import My_YOLO as my_yolov5s
from yolov5l import My_YOLO as my_yolov5l
from yolov5m import My_YOLO as my_yolov5m
from yolov5x import My_YOLO as my_yolov5x

device = 'cuda' if torch.cuda.is_available() else 'cpu'
parser = argparse.ArgumentParser()
parser.add_argument('--epochs', type=int, default=300)
parser.add_argument('--imgroot', default='/home/wangbo/Desktop/data/datasets/VOCdevkit/VOC2012/JPEGImages', help='image directory path')
parser.add_argument('--labroot', default='/home/wangbo/Desktop/data/datasets/VOCdevkit/VOC2012/Annotations', help='label directory path')
parser.add_argument('--train_txt', default='/home/wangbo/Desktop/data/datasets/VOCdevkit/VOC2012/ImageSets/Main/trainval.txt', help='train.txt')
parser.add_argument('--test_txt', default='/home/wangbo/Desktop/data/datasets/VOCdevkit/VOC2012/ImageSets/Main/val.txt', help='test.txt')
parser.add_argument('--labels_txt', default='pascal_voc.names', help='labels.txt')
parser.add_argument('--save_model', type=str, required=True, help='save weights dirname')
parser.add_argument('--resume', action='store_true', help='resume training from last.pt')
parser.add_argument('--augment', action='store_true', help='data augment')
parser.add_argument('--imgsize', type=int, default=640, help='image size')
parser.add_argument('--input_norm', action='store_true', help='Input Normliaze')
parser.add_argument('--batchsize', type=int, default=16, help='Batch size')
parser.add_argument('--pretrained', type=str, default='', help='Pretrained parameter file')
parser.add_argument("--savepth_interval", default=10, type=int)
parser.add_argument("--test_interval", default=5, type=int, help='if not test , set <=0')
parser.add_argument('--net_type', default='yolov5s', choices=['yolov5s', 'yolov5l', 'yolov5m', 'yolov5x'])
parser.add_argument("--optimizer", default='adam', choices=['adam', 'sgd'])
parser.add_argument('--lr', type=float, default=1e-3, help='learning rate')
parser.add_argument('--momentum', default=0.9, type=float, help='Momentum value for optim sgd')
parser.add_argument("--weight_decay", default=1e-4, type=float)
parser.add_argument("--iou_thres", type=float, default=0.5, help="iou threshold required to qualify as detected")
parser.add_argument("--conf_thres", type=float, default=0.5, help="object confidence threshold on test dataset")
parser.add_argument("--nms_thres", type=float, default=0.5, help="iou thresshold for non-maximum suppression")
parser.add_argument("--accord_type", default='train_loss', choices=['train_loss', 'test_map', 'test_loss'])
parser.add_argument('--plot_loss', action='store_true', help='plot train loss curve')
parser.add_argument('--write_excel', action='store_true', help='write train logs to excel')
parser.add_argument('--keep_ratio', action='store_true', help='resize image keep height and width ratio')
args = parser.parse_args()
print(args)

if __name__=="__main__":
    trainDataSet = MyDataset_xml(args.imgroot, args.labroot, args.train_txt, args.labels_txt, image_size=args.imgsize, input_norm=args.input_norm, augment=args.augment)
    # trainDataSet = MyDataset_json(args.imgroot, args.labroot, args.train_txt, args.labels_txt, image_size=args.imgsize, input_norm=args.input_norm, augment=args.augment)
    print("Train dataset size: {}".format(len(trainDataSet)))
    dataloaderTrain = DataLoader(trainDataSet,
                                batch_size=args.batchsize,
                                shuffle=True,
                                num_workers=0,
                                pin_memory=True,
                                collate_fn=trainDataSet.collate_fn)

    testDataSet = MyDataset_xml(args.imgroot, args.labroot, args.test_txt, args.labels_txt, image_size=args.imgsize, input_norm=args.input_norm)
    # testDataSet = MyDataset_json(args.imgroot, args.labroot, args.test_txt, args.labels_txt, image_size=args.imgsize, input_norm=args.input_norm, train=False)
    print("validation dataset size: {}".format(len(testDataSet)))
    dataloaderTest = DataLoader(testDataSet,
                                 batch_size=args.batchsize,
                                 shuffle=False,
                                 num_workers=0,
                                 pin_memory=True,
                                 collate_fn=testDataSet.collate_fn)

    anchors = [[10, 13, 16, 30, 33, 23], [30, 61, 62, 45, 59, 119], [116, 90, 156, 198, 373, 326]]
    # net = My_YOLO(trainDataSet.num_classes, anchors=anchors)
    if args.net_type=='yolov5s':
        net = my_yolov5s(trainDataSet.num_classes, anchors=anchors, training=True)
    elif args.net_type=='yolov5l':
        net = my_yolov5l(trainDataSet.num_classes, anchors=anchors, training=True)
    elif args.net_type=='yolov5m':
        net = my_yolov5m(trainDataSet.num_classes, anchors=anchors, training=True)
    else:
        net = my_yolov5x(trainDataSet.num_classes, anchors=anchors, training=True)
    net.to(device)
    if len(args.pretrained) > 0:
        net.load_state_dict(torch.load(args.pretrained, map_location=device))
    # else:
    #     net.apply(weights_init_normal)

    train_dir = os.path.join(os.getcwd(), args.save_model)
    if not args.resume:
        if os.path.exists(train_dir):
            shutil.rmtree(train_dir)
        os.makedirs(train_dir)

    starting_step, accord_type = 0, args.accord_type
    best_fitness = 0 if accord_type == 'test_map' else float('inf')
    nb, train_logs = len(dataloaderTrain), {}
    if args.optimizer == 'adam':
        optimizer = Adam(net.parameters(), lr=args.lr)
    else:
        optimizer = SGD(net.parameters(), lr=args.lr, momentum=args.momentum, weight_decay=args.weight_decay)

    scheduler = lr_scheduler.MultiStepLR(optimizer, milestones=[round(args.epochs * x) for x in
                                                                [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]])
    # scheduler = lr_scheduler.CosineAnnealingLr(optimizer, 50)

    filenameCheckpoint = train_dir + '/last.pth'  ###load pretrain model
    if args.resume and os.path.exists(filenameCheckpoint):
        checkpoint = torch.load(filenameCheckpoint, map_location=device)
        starting_step = checkpoint['step']
        # scheduler.last_epoch = starting_step
        # checkpoint['optimizer']['param_groups'][0]['lr'] = 0.00000519
        net.load_state_dict(checkpoint['state_dict'])
        optimizer.load_state_dict(checkpoint['optimizer'])
        scheduler.load_state_dict(checkpoint['scheduler_state'])
        accord_type = checkpoint['accord_type']
        best_fitness = checkpoint['best_fitness']
        train_logs = checkpoint['train_logs']
        print("=> Loaded checkpoint at step {})".format(checkpoint['step']))
        del checkpoint
    if args.test_interval > 0:
        test_ind = [i for i in range(1, args.epochs) if i % args.test_interval == 0]
        if test_ind[-1] != args.epochs - 1:
            test_ind.append(args.epochs - 1)
        test_ind = tuple(test_ind)
    else:
        test_ind = ()

    net.hyp = hyp
    net.hyp['cls'] *= trainDataSet.num_classes / 80.
    net.nc = trainDataSet.num_classes
    net.gr = 1.0
    for epoch in range(starting_step, args.epochs):
        net.train()
        pbar = tqdm(enumerate(dataloaderTrain), total=nb)  # progress bar
        all_loss = 0
        start = time.time()
        for i, (inputs,labels) in pbar:
            inputs = inputs.permute(0, 3, 1, 2).to(device)
            labels = labels.to(device)
            pred = net(inputs)
            loss, loss_items = compute_loss(pred, labels, net)
            loss.backward()
            optimizer.step()
            optimizer.zero_grad()
            all_loss += loss.item()
            pbar.set_description('epoch' + str(epoch))
        end = time.time()
        scheduler.step(epoch=epoch)
        train_logs[epoch] = {'train_loss': all_loss / nb, 'waste_time': end - start}
        if accord_type == 'train_loss':
            if train_logs[epoch]['train_loss'] < best_fitness:
                best_fitness = train_logs[epoch]['train_loss']
                torch.save(net.state_dict(), os.path.join(train_dir, 'best.pth'))

        if epoch in test_ind:
            results = evaluate(net, dataloaderTest, args.iou_thres, args.conf_thres, args.nms_thres, args.imgsize)
            if isinstance(results, (tuple,list)):
                precision, recall, AP, f1, ap_class, map, test_loss = results
                if accord_type == 'test_loss':
                    if test_loss < best_fitness:
                        best_fitness = test_loss
                        torch.save(net.state_dict(), os.path.join(train_dir, 'best.pth'))
                elif accord_type == 'test_map':
                    if map > best_fitness:
                        best_fitness = map
                        torch.save(net.state_dict(), os.path.join(train_dir, 'best.pth'))
                train_logs[epoch].update({'test_precision':precision.mean(), 'test_recall':recall.mean(), 'test_mAP':map, 'test_f1':f1.mean(), 'test_loss':test_loss, 'best_fitness':best_fitness})
            else:
                if accord_type == 'test_loss':
                    if results < best_fitness:
                        best_fitness = results
                        torch.save(net.state_dict(), os.path.join(train_dir, 'best.pth'))
                train_logs[epoch]['test_loss'] = results
        print('epoch['+str(epoch)+']:', train_logs[epoch])
        if epoch % args.savepth_interval == 0 and epoch > 0:
            torch.save(net.state_dict(), os.path.join(train_dir, 'epoch' + str(epoch) + '.pth'))

        lastinfo = {'step': epoch, 'accord_type': accord_type,
                    'best_fitness': best_fitness,
                    'state_dict': net.state_dict(),
                    'optimizer': optimizer.state_dict(),
                    'scheduler_state': scheduler.state_dict(),
                    'train_logs': train_logs}
        torch.save(lastinfo, filenameCheckpoint)
        torch.cuda.empty_cache()
    np.save('finish.npy', np.array([1, 1, 1]))

    if args.plot_loss:
        ind = [i for i in range(args.epochs) if 'train_loss' in train_logs[i]]
        x_axis = np.asarray(ind)
        y_axis = [train_logs[i]['train_loss'] for i in ind]
        y_axis = np.asarray(y_axis)
        plt.plot(x_axis, y_axis, color='red', label='train loss')

        ind = [i for i in range(args.epochs) if 'test_mAP' in train_logs[i]]
        x_axis = np.asarray(ind)
        y_axis = [train_logs[i]['test_mAP'] for i in ind]
        y_axis = np.asarray(y_axis)
        plt.plot(x_axis, y_axis, color='green', label='test mAP')

        ind = [i for i in range(args.epochs) if 'test_precision' in train_logs[i]]
        x_axis = np.asarray(ind)
        y_axis = [train_logs[i]['test_precision'] for i in ind]
        y_axis = np.asarray(y_axis)
        plt.plot(x_axis, y_axis, color='blue', label='test precision')

        ind = [i for i in range(args.epochs) if 'test_recall' in train_logs[i]]
        x_axis = np.asarray(ind)
        y_axis = [train_logs[i]['test_recall'] for i in ind]
        y_axis = np.asarray(y_axis)
        plt.plot(x_axis, y_axis, color='yellow', label='test recall')

        ind = [i for i in range(args.epochs) if 'test_loss' in train_logs[i]]
        x_axis = np.asarray(ind)
        y_axis = [train_logs[i]['test_loss'] for i in ind]
        y_axis = np.asarray(y_axis)
        plt.plot(np.asarray(x_axis), np.asarray(y_axis), color='black', label='test loss')

        plt.legend()
        plt.xlabel('epoch')
        plt.ylabel('train_log')
        plt.savefig(os.path.join(train_dir, 'train_log.jpg'), dpi=300)

    if args.write_excel:
        key_names = set()
        for i in range(args.epochs):
            names = set(train_logs[i].keys())
            key_names = key_names.union(names)
        key_names = list(key_names)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.cell(1, 1, 'epoch')
        for i in range(len(key_names)):
            ws.cell(1, i+2, key_names[i])
        for i in range(args.epochs):
            ws.cell(i+2, 1, str(i))
            for j, name in enumerate(key_names):
                s = str(train_logs[i][name]) if name in train_logs[i] else ''
                ws.cell(i + 2, j+2, s)

        wb.save(os.path.join(train_dir, 'train_log.xlsx'))